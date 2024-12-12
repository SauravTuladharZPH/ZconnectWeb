from openpyxl import load_workbook
from robot.api.deco import keyword
import imaplib
import email
import re
from bs4 import BeautifulSoup
from email.header import decode_header

from selenium import webdriver
from robot.libraries.BuiltIn import BuiltIn

menu_items = []


# Function to read a single cell's value
def read_excel_data(file_path, sheet_name, cell):
    # Load the workbook and sheet
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    # Get the value of the specified cell
    cell_value = sheet[cell].value
    return cell_value


# Function to read a range of data (like a table of menu data)
def read_excel_range(file_path, sheet_name, start_cell, end_cell):
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]
    rows = sheet[start_cell:end_cell]
    data = []
    for row in rows:
        menu_name = row[0].value.strip() if row[0].value else None
        expected_heading = row[1].value.strip() if row[1].value else None
        if menu_name and expected_heading:
            data.append((menu_name, expected_heading))
    workbook.close()
    # Log the structure of the data to ensure it's flat
    print(f"Data: {data}")
    return data  # Should return a flat list of tuples


@keyword
def read_excel_registration_data(file_path, sheet_name):
    # Open the workbook and sheet
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    # Get headers from the first row
    headers = [cell.value if cell.value is not None else "" for cell in sheet[1]]
    # Initialize an empty list to hold all row data
    data = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=len(headers)):
        # Extract cell values for the current row, replacing None with empty string
        row_values = [cell.value if cell.value is not None else "" for cell in row]
        # Combine headers and row values into a dictionary
        row_data = dict(zip(headers, row_values))
        # Append to the data list if row is not empty
        if any(row_values):
            data.append(row_data)
    workbook.close()
    return data


def fetch_otp_from_gmail(email_user, email_password, sender_email, subject):
    try:
        # Connect to Gmail's IMAP server
        imap = imaplib.IMAP4_SSL("imap.gmail.com")

        # Login to your account
        imap.login(email_user, email_password)

        # Select the inbox folder
        imap.select("INBOX")

        # Search for unread emails from the sender with the specific subject
        status, messages = imap.search(
            None, f'(FROM "{sender_email}" SUBJECT "{subject}" UNSEEN)'
        )
        email_ids = messages[0].split()

        # If no unread emails are found, return None
        if not email_ids:
            print("No unread emails found matching the criteria.")
            return None

        # Fetch only the most recent unread email
        latest_email_id = email_ids[-1]
        status, msg_data = imap.fetch(latest_email_id, "(RFC822)")

        body = None  # Initialize variable for email body

        for response_part in msg_data:
            if isinstance(response_part, tuple):
                # Parse the email content
                msg = email.message_from_bytes(response_part[1])
                print(f"Processing email with subject: {msg['Subject']}")

                # Check if the email is multipart
                if msg.is_multipart():
                    for part in msg.walk():
                        content_type = part.get_content_type()
                        if content_type == "text/plain":
                            body = part.get_payload(decode=True).decode()
                            break
                        elif content_type == "text/html":
                            # Extract and clean HTML content
                            raw_html = part.get_payload(decode=True).decode()
                            body = BeautifulSoup(raw_html, "html.parser").get_text()
                            break
                else:
                    # For non-multipart emails
                    raw_html = msg.get_payload(decode=True).decode()
                    body = BeautifulSoup(raw_html, "html.parser").get_text()

        if not body:
            print("Failed to retrieve the email body.")
            return None

        # Normalize the email body
        body = " ".join(body.split())
        print(f"Normalized Email Body:\n{body}\n")

        # Extract OTP using regex
        otp_pattern = r"Your OTP code is: (\d{6})"
        match = re.search(otp_pattern, body)
        if match:
            otp = match.group(1)
            print(f"Extracted OTP: {otp}")
            return otp

        print("No OTP found in the email.")
        return None

    except Exception as e:
        print(f"An error occurred: {e}")
        return None

    finally:
        # Close the connection and logout
        try:
            imap.close()
            imap.logout()
        except:
            pass


def validate_username_from_gmail(email_user, email_password, sender_email, subject, username, group, logincode,
                                 email_address):
    try:
        # Connect to Gmail's IMAP server
        imap = imaplib.IMAP4_SSL("imap.gmail.com")
        # Login to your account
        imap.login(email_user, email_password)
        # Select the mailbox
        imap.select("INBOX")
        # Search for emails from the sender with the subject
        status, messages = imap.search(
            None, f'FROM "{sender_email}" SUBJECT "{subject}"'
        )
        # Check if emails are found
        email_ids = messages[0].split()
        if not email_ids:
            print("No emails found matching the criteria.")
            return None
        # Fetch the latest email
        latest_email_id = email_ids[-1]
        status, msg_data = imap.fetch(latest_email_id, "(RFC822)")
        body = None  # Initialize body to avoid uninitialized variable error
        for response_part in msg_data:
            if isinstance(response_part, tuple):
                # Parse the email content
                msg = email.message_from_bytes(response_part[1])
                print(f"Processing email with subject: {msg['Subject']}")
                # Check email parts
                if msg.is_multipart():
                    for part in msg.walk():
                        content_type = part.get_content_type()
                        print(f"Found part with content type: {content_type}")
                        if content_type == "text/plain":
                            body = part.get_payload(decode=True).decode()
                            break
                        elif content_type == "text/html":
                            # Fallback to HTML if plain text is not available
                            body = part.get_payload(decode=True).decode()
                            break
                else:
                    # If the email is not multipart
                    body = msg.get_payload(decode=True).decode()
        if not body:
            print("Failed to retrieve the email body.")
            return None
        # Print the body for inspection (optional)
        print(f"Email Body:\n{body}\n")
        # Validate the content in the email body with parameters
        required_fields = {
            'Username': username,
            'Group': group,
            'Login Code': logincode,
            'Email Address': email_address
        }
        for key, value in required_fields.items():
            if f"{key}: {value}" not in body:
                print(f"Missing or incorrect {key}. Expected '{key}: {value}'.")
                return None
            else:
                print(f"Validated {key}: {value}")
        print("Email content validation passed.")
        return True  # Return success if everything is validated
    except Exception as e:
        print(f"An error occurred: {e}")
        return None
    finally:
        # Close the connection and logout
        imap.close()
        imap.logout()


def normalize_text(text):
    """
    Normalizes text by removing extra whitespace, line breaks, and converting to lowercase.
    """
    return " ".join(text.split()).strip().lower()


def validate_locateInformation_from_gmail(email_user, email_password, sender_email, subject, clientname, group,
                                          logincode, username):
    """
    Validates the presence of specific information in an email body retrieved from Gmail.
    """
    try:
        # Connect to Gmail's IMAP server
        imap = imaplib.IMAP4_SSL("imap.gmail.com")

        # Login to your Gmail account
        imap.login(email_user, email_password)

        # Select the inbox folder
        imap.select("INBOX")

        # Search for the latest email matching the sender and subject
        status, messages = imap.search(None, f'FROM "{sender_email}" SUBJECT "{subject}"')
        email_ids = messages[0].split()
        if not email_ids:
            print("No emails found matching the criteria.")
            return None

        # Fetch the latest email
        latest_email_id = email_ids[-1]
        status, msg_data = imap.fetch(latest_email_id, "(RFC822)")

        body = None  # Initialize variable to store the email body

        for response_part in msg_data:
            if isinstance(response_part, tuple):
                # Parse the email content
                msg = email.message_from_bytes(response_part[1])
                print(f"Processing email with subject: {msg['Subject']}")

                # Check if the email is multipart
                if msg.is_multipart():
                    for part in msg.walk():
                        content_type = part.get_content_type()
                        if content_type == "text/plain":
                            body = part.get_payload(decode=True).decode()
                            break
                        elif content_type == "text/html":
                            raw_html = part.get_payload(decode=True).decode()
                            body = BeautifulSoup(raw_html, "html.parser").get_text()
                            break
                else:
                    raw_html = msg.get_payload(decode=True).decode()
                    body = BeautifulSoup(raw_html, "html.parser").get_text()

        if not body:
            print("Failed to retrieve the email body.")
            return None

        # Normalize the email body for easier validation
        body = normalize_text(body)

        print(f"Normalized Email Body:\n{body}\n")  # Debugging

        # Define the expected content, normalized for comparison
        required_fields = {
            "client name: zakipoint demo b": clientname.lower(),
            "group name: headquarter": group.lower(),
            "login group id: zph1": logincode.lower(),
            "username: sauravzakipoint": username.lower(),
        }

        # Validate each required field
        for key, value in required_fields.items():
            if key not in body:
                print(f"Missing or incorrect {key}. Expected '{value}'.")
                return None
            else:
                print(f"Validated {key}.")

        print("Email content validation passed.")
        return True

    except Exception as e:
        print(f"An error occurred: {e}")
        return None

    finally:
        # Ensure cleanup of resources
        try:
            imap.close()
            imap.logout()
        except:
            pass


@keyword("Read HomepageMenu From Excel")
def read_homepagemenu_from_excel(file_path, sheet_name):
    try:
        workbook = load_workbook(file_path)
        sheet = workbook[sheet_name]
        # Collect menu items from the first column (ignoring empty cells)
        menu_items = [row[0].value.strip() for row in sheet.iter_rows(min_row=2, max_col=1) if row[0].value]
        return menu_items
    except Exception as e:
        raise Exception(f"Error reading Excel file: {e}")


@keyword("Read Homepage Menu Data From Excel")
def read_homepage_menu_data_from_excel(file_path, sheet_name):
    try:
        workbook = load_workbook(file_path, data_only=True)  # Ensure data is read correctly
        sheet = workbook[sheet_name]
        menu_data = []

        # Ensure the header row is skipped (assuming row 1 is the header)
        for row in sheet.iter_rows(min_row=2, max_col=3):
            # Extract values from row
            menu_name = row[0].value if row[0].value else ""
            expected_heading = row[1].value if row[1].value else ""
            url = row[2].value if row[2].value else ""

            # Only add non-empty rows
            if menu_name:
                menu_data.append({
                    "menu_name": menu_name,
                    "expected_heading": expected_heading,
                    "url": url,
                })

        return menu_data
    except Exception as e:
        raise Exception(f"Error reading Excel file: {e}")


@keyword("Read Cost Estimator Provider Category and Specialities Test Data")
def read_provider_category_and_specialities_data(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = {}
    # Iterate through the rows, starting from the second row
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming headers are in the first row
        category = row[4]  # Column E (zero-indexed: 4th column)
        specialities = row[5]  # Column F (zero-indexed: 5th column)
        if category and specialities:
            # Split the specialties by comma and strip extra spaces
            data[category] = [s.strip() for s in specialities.split(",")]
    return data